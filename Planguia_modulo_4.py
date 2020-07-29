from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import Planguia_funcoes
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Protection
from openpyxl.styles import Font, numbers
import requests
import json
import win32com.client as win32


class EstimativaCustos:
    def __init__(self, arquivo, aba_leitura, aba_destino, cambio):
        self.arquivo = arquivo
        self.aba_leitura = aba_leitura
        self.aba_destino = aba_destino
        self.cambio = cambio

    def listar_taxadiaria(self):
        w = Planguia_funcoes.openr(self.arquivo, self.aba_leitura)
        dict_taxas = {}
        lista_dados = []
        numero_linhas = w[1].max_row
        for linhas in range(4, numero_linhas + 1):
            equipamento = w[1].cell(row=linhas, column=1).value
            nome = w[1].cell(row=linhas, column=2).value
            tipo = w[1].cell(row=linhas, column=3).value
            dolar = float(w[1].cell(row=linhas, column=4).value)
            real = float(w[1].cell(row=linhas, column=5).value)
            reajuste = float(w[1].cell(row=linhas, column=6).value)
            valor_taxa = round(((real * reajuste) / self.cambio) + dolar, 2)
            lista_dados.append(nome)
            lista_dados.append(tipo)
            lista_dados.append(valor_taxa)
            dict_taxas[equipamento] = lista_dados
            dict_taxas.copy()
            lista_dados = []
        return dict_taxas

    def listar_estimativa(self):
        w1 = Planguia_funcoes.openr(self.arquivo, 'Previa')
        lista_geral = []
        lista_embarc = []
        numero_linha = w1[1].max_row
        for linha in range(2, numero_linha + 1):
            equip_previa = w1[1].cell(row=linha, column=3).value
            embarc_previa = w1[1].cell(row=linha, column=1).value
            dias_petro = round(w1[1].cell(row=linha, column=11).value, 3)
            dias_pblog = round(w1[1].cell(row=linha, column=14).value, 3)
            lista_embarc.append(equip_previa)
            lista_embarc.append(embarc_previa)
            lista_embarc.append(dias_petro)
            lista_embarc.append(dias_pblog)
            lista_geral.append(lista_embarc)
            lista_embarc = []
        return lista_geral

    def catalogar_dados(self):
        planilha_estimativa = self.listar_estimativa()
        dados_taxa = self.listar_taxadiaria()
        planilha = Planguia_funcoes.openr(self.arquivo, self.aba_destino)
        linha = 3
        for listas in range(0, len(planilha_estimativa)):
            dados = planilha_estimativa[listas]

            planilha[1].cell(row=linha, column=1).value = dados[0]
            planilha[1].cell(row=linha, column=1).alignment = Alignment(horizontal='center')
            equipamento1 = planilha[1].cell(row=linha, column=1).value

            planilha[1].cell(row=linha, column=2).value = dados[1].upper()
            planilha[1].cell(row=linha, column=2).alignment = Alignment(horizontal='center')

            lista_resultados = dados_taxa[equipamento1]
            planilha[1].cell(row=linha, column=3).value = lista_resultados[1]
            planilha[1].cell(row=linha, column=3).alignment = Alignment(horizontal='center')

            taxa_diaria = lista_resultados[2]
            indice_petro = dados[2]
            indice_pblog = dados[3]

            valor_petro = round(taxa_diaria * indice_petro, 2)
            planilha[1].cell(row=linha, column=4).value = Planguia_funcoes.converter_moeda(valor_petro)
            planilha[1].cell(row=linha, column=4).alignment = Alignment(horizontal='center')

            valor_pblog = round(taxa_diaria * indice_pblog, 2)
            planilha[1].cell(row=linha, column=5).value = Planguia_funcoes.converter_moeda(valor_pblog)
            planilha[1].cell(row=linha, column=5).alignment = Alignment(horizontal='center')

            somatorio = round(valor_petro + valor_pblog, 2)
            planilha[1].cell(row=linha, column=6).value = Planguia_funcoes.converter_moeda(somatorio)
            planilha[1].cell(row=linha, column=6).alignment = Alignment(horizontal='center')
            planilha[1].cell(row=linha, column=6).font = Font(bold=True)

            linha = linha + 1
        Planguia_funcoes.closer('Projeto_planilha_Guia_Medição.xlsx', planilha[0])

    def formatar(self):
        planilha = Planguia_funcoes.openr(self.arquivo, self.aba_destino)
        planilha[1].cell(row=1, column=1).value = 'Cambio: ' + str(self.cambio)
        cabecalho = ['Equipamento', 'Embarcação', 'Tipo', 'PB1', 'PB2', 'Total']

        for dados in range(1, len(cabecalho) + 1):
            planilha[1].cell(row=2, column=dados).value = cabecalho[dados - 1]
            cores_verdec = PatternFill(fill_type='solid', start_color='c6ffb3', end_color='c6ffb3')
            planilha[1].cell(row=2, column=dados).fill = cores_verdec
            planilha[1].cell(row=2, column=dados).alignment = Alignment(horizontal='center')
            planilha[1].cell(row=2, column=dados).font = Font(bold=True)

        planilha[1].cell(row=1, column=1).font = Font(bold=True)
        planilha[1].cell(row=1, column=1).alignment = Alignment(horizontal='center')

        numero_linha = planilha[1].max_row
        for i in range(3, numero_linha+1):
            regiao1 = planilha[1]['D{}'.format(i)]
            regiao2 = planilha[1]['E{}'.format(i)]
            regiao3 = planilha[1]['F{}'.format(i)]
            thin = Side(border_style="thin", color="000000")
            regiao1.border = Border(top=None, left=thin, right=thin, bottom=None)
            regiao2.border = Border(top=None, left=thin, right=thin, bottom=None)
            regiao3.border = Border(top=None, left=thin, right=thin, bottom=None)

        lista_celula = ['D', 'E', 'F']
        for indice in lista_celula:
            celula = planilha[1]['{}2'.format(indice)]
            thin1 = Side(border_style="thin", color="000000")
            celula.border = Border(top=thin1, left=thin1, right=thin1, bottom=thin1)

        lista_celula1 = ['A', 'B', 'C']
        for indice1 in lista_celula1:
            celula = planilha[1]['{}2'.format(indice1)]
            thin2 = Side(border_style="thin", color="000000")
            celula.border = Border(top=thin2, left=None, right=None, bottom=thin2)

        lista_celula2 = ['A', 'B', 'C', 'D', 'E', 'F']
        numero_linha1 = planilha[1].max_row
        for indice2 in lista_celula2:
            celula = planilha[1]['{}{}'.format(indice2, numero_linha1)]
            thin2 = Side(border_style="thin", color="000000")
            celula.border = Border(top=None, left=None, right=None, bottom=thin2)

        lista_celula2 = ['D', 'E', 'F']
        numero_linha2 = planilha[1].max_row
        for indice2 in lista_celula2:
            celula = planilha[1]['{}{}'.format(indice2, numero_linha2)]
            thin3 = Side(border_style="thin", color="000000")
            celula.border = Border(top=None, left=thin3, right=thin3, bottom=thin3)

        Planguia_funcoes.closer('Projeto_planilha_Guia_Medição.xlsx', planilha[0])


def iniciar_4():
    valor_cambio = Planguia_funcoes.importar_cambio()
    custo = EstimativaCustos('Projeto_planilha_Guia_Medição.xlsx', 'Taxa', 'Estimativa Custo', valor_cambio)
    custo.catalogar_dados()
    custo.formatar()
    Planguia_funcoes.ajustar_colunas(r'C:\Users\(chave)\Desktop\Python\projetos\Projeto_planilha_Guia_Medição.xlsx', 'Estimativa Custo')
