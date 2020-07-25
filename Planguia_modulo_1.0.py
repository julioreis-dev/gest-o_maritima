import Planguia_funcoes
import pandas as pd
import xlsxwriter
import time
from openpyxl import load_workbook
import Planguia_modulo_2
import Planguia_modulo_3
import Planguia_modulo_4
import Planguia_inoperancia
import win32com.client as win32

'''Esta classe molda a planilha inicial de dados referentes a movimentação da frota maritima.'''


class PlanilhaInicial:
    def __init__(self, arquivo, aba, aba1):
        self.arquivo = arquivo
        self.aba = aba
        self.aba1 = aba1

    def documentar_equipamento(self):
        """ Prepara um dicionário:ex:{ICJ: equipamento}."""
        dados_arquivo = Planguia_funcoes.openr(self.arquivo, self.aba)
        contador_linha = dados_arquivo[1].max_row
        dict_equipamento = {}
        for n in range(2, contador_linha + 1):
            numero_icj = dados_arquivo[1].cell(row=n, column=1).value
            dict_equipamento[numero_icj] = dados_arquivo[1].cell(row=n, column=2).value
            dict_equipamento.copy()
        return dict_equipamento

    def preparar_tupla(self):
        """Prepara uma Tupla com os dados relevantes de medição da aba com as informações recebidos."""
        dados_arquivo = Planguia_funcoes.openr(self.arquivo, self.aba1)
        contador = dados_arquivo[1].max_row
        lista_dados = []
        for linha in range(2, contador + 1):
            embarcacao = dados_arquivo[1].cell(row=linha, column=1).value
            icj = dados_arquivo[1].cell(row=linha, column=2).value
            tipo = dados_arquivo[1].cell(row=linha, column=3).value
            regional = dados_arquivo[1].cell(row=linha, column=4).value
            regional_1 = dados_arquivo[1].cell(row=linha, column=9).value
            dias_operacao = dados_arquivo[1].cell(row=linha, column=21).value
            tupla_dados = (embarcacao, icj, tipo, regional, regional_1, round(dias_operacao, 3))
            lista_dados.append(tupla_dados)
        return lista_dados

    def organizar(self):
        """Organiza a lista de dados de medição em ordem alfabética e escreve na planilha destino."""
        dados_arquivo = Planguia_funcoes.openr(self.arquivo, self.aba1)
        dados_arquivo[0].create_sheet('Previa')
        nova_aba = dados_arquivo[0].create_sheet('Base Dados')
        organizado = sorted(self.preparar_tupla())
        dict_dados = self.documentar_equipamento()
        linha_preenchida = 2
        for t in organizado:
            nova_aba.cell(row=linha_preenchida, column=2).value = t[0]
            nova_aba.cell(row=linha_preenchida, column=3).value = t[1]
            nova_aba.cell(row=linha_preenchida, column=4).value = t[2]
            nova_aba.cell(row=linha_preenchida, column=5).value = t[3]
            nova_aba.cell(row=linha_preenchida, column=6).value = t[4]
            nova_aba.cell(row=linha_preenchida, column=7).value = t[5]
            icj_extraido = t[1]
            equip = dict_dados[icj_extraido]
            nova_aba.cell(row=linha_preenchida, column=1).value = equip
            linha_preenchida = linha_preenchida + 1

        cabecalho = ['Equipamento', 'Embarcação', 'ICJ', 'Tipo', 'Regional', 'Regional1', 'Dias']
        for v1 in range(1, len(cabecalho) + 1):
            nova_aba.cell(row=1, column=v1).value = cabecalho[v1 - 1]
        Planguia_funcoes.closer('Projeto_planilha_Guia_Medição.xlsx', dados_arquivo[0])

    def ajustar_gopi(self):
        gopi = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição.xlsx', 'GOPI')
        contador_linha = gopi[1].max_row
        dict_relevante = self.documentar_equipamento()
        for linha_gopi in range(2, contador_linha + 1):
            icj_gopi = gopi[1].cell(row=linha_gopi, column=2).value
            equipamento = dict_relevante[icj_gopi]
            gopi[1].cell(row=linha_gopi, column=2).value = equipamento
        Planguia_funcoes.closer('Projeto_planilha_Guia_Medição.xlsx', gopi[0])

    @staticmethod
    def agregar_valores():
        Planguia_funcoes.agregar_dados()
        arquivo_excel = r'C:\Users\(chave)\Desktop\Python\projetos\Projeto_planilha_Guia_Medição_2020_1.xlsx'
        planilha = pd.read_excel(arquivo_excel, sheet_name='Base Dados')

        planilha_nova = planilha[['Embarcação', 'Regional1', 'Equipamento', 'Regional', 'Tipo', 'Dias']]
        destino = pd.ExcelWriter('Projeto_planilha_Guia_Medição_2020_1.xlsx')
        planilha_nova.to_excel(destino, 'Base Dados', index=False)
        destino.save()

    @staticmethod
    def ajustar_celulas():
        wb = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020_1.xlsx', 'Base Dados')
        wb2 = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição.xlsx', 'Previa')
        contar_linha = wb[1].max_row
        for n in range(2, contar_linha + 1):
            for t in range(1, 4):
                atual = wb[1].cell(row=n, column=t).value
                if atual is None:
                    anterior = wb[1].cell(row=n - 1, column=t).value
                    wb[1].cell(row=n, column=t).value = anterior

        for linha in range(1, contar_linha + 1):
            for coluna in range(1, 7):
                wb2[1].cell(row=linha, column=coluna).value = wb[1].cell(row=linha, column=coluna).value
        Planguia_funcoes.closer('Projeto_planilha_Guia_Medição.xlsx', wb2[0])
        wb[0].close()


def opcao1():
    planguia = PlanilhaInicial('Planilha Guia_dados.xlsx', 'Info Contrato', 'GOPI')
    func = Planguia_funcoes
    planguia.organizar()
    planguia.agregar_valores()
    planguia.ajustar_celulas()
    planguia.ajustar_gopi()
    Planguia_modulo_2.iniciar_2()
    func.ajustar_cabecalho()
    func.ajustar_colunas(r'C:\Users\ay4m\Desktop\Python\projetos\Projeto_planilha_Guia_Medição.xlsx', 'Previa')


def opcao2():
    func = Planguia_funcoes
    func.agrupar_inoperancias()
    func.realocar_inoperancias()
    func.calcular()
    Planguia_modulo_4.iniciar_4()


def opcao3():
    func = Planguia_funcoes
    func.deletar('Projeto_planilha_Guia_Medição.xlsx', 'Medição')
    func.calcular()
    func.transferir_dados()
    func.ajustar_colunas(r'C:\Users\(chave)\Desktop\Python\projetos\Projeto_planilha_Guia_Medição.xlsx', 'Medição')


def opcao4():
    opcao1()
    opcao2()
    opcao3()


def opcao5():
    Planguia_modulo_3.iniciar_3()


def main():
    func = Planguia_funcoes
    arg_pass = True
    while arg_pass:
        pergunta = input('################################################'
                         '\nTipos de opções disponíveis nesta aplicação:'
                         '\nDigite 1 --> Emitir prévia da planilha guia.'
                         '\nDigite 2 --> Computar indices de inoperâncias.'
                         '\nDigite 3 --> Preparar versão final da planilha guia.'
                         '\nDigite 4 --> (Opcional) - Realizar o processo completo de elaboração da Planilha Guia.'
                         '\nDigite 5 --> (Opcional) - Enviar planilha guia para os gerentes e fiscais de contrato.'
                         '\nDigite 0 --> Sair.'
                         '\n################################################'
                         '\nPrezado usuário, escolha uma opção?')
        if pergunta.isdigit():
            pergunta = int(pergunta)
            if pergunta == 1:
                print('Previa da planilha guia sendo realizada.\nProcessando................')
                opcao1()
                func.mostrar_desempenho(5, 0)
                func.mostrar_desempenho(0, 1)
            elif pergunta == 2:
                print('Analisando inoperâncias da frota.\nProcessando................')
                opcao2()
                func.mostrar_desempenho(3, 2)
            elif pergunta == 3:
                opcao3()
                func.mostrar_desempenho(0, 4)
            elif pergunta == 4:
                print('Realizando o processo completo de elaboração da planilha guia.\nProcessando................')
                opcao4()
                func.mostrar_desempenho(0, 4)
            elif pergunta == 5:
                print('Preparando email para envio.\nProcessando................')
                opcao5()
            elif pergunta == 0:
                print('Finalizando................')
                arg_pass = False
            else:
                print('\nOpção inválida, digite novamente de acordo com as opções disponíveis!!!')
            time.sleep(3)
        else:
            print('\nPrezado usuário, tente novamente digitando um numero válido!!!\n')
            time.sleep(3)
    print('\nPrezado usuário, aplicação encerrada com sucesso!!!\n')


if __name__ == '__main__':
    main()
