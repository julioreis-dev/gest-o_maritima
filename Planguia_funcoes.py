import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import xlrd
from openpyxl.styles import PatternFill, Border, Side, Protection
import time
from openpyxl.styles import Font, numbers
from openpyxl.styles import colors
import win32com.client as win32
import Planguia_modulo_2
import time
from openpyxl.styles import Alignment
import requests
import json
from babel.numbers import format_currency

'''Funções de apoio que são necessárias para a elaboração da planilha guia.'''


def openr(arquivo, aba):
    lista = []
    wb = load_workbook(filename=arquivo)
    aba_dados = wb[aba]
    lista.append(wb)
    lista.append(aba_dados)
    return lista


def closer(nome_arquivo, wb):
    wb.save(filename=nome_arquivo)
    wb.close()


def contador_linhas(endereco, abaplanilha):
    planilha_atual = pd.read_excel(endereco, sheet_name=abaplanilha)
    valor = planilha_atual.shape[0]
    return valor + 1


def contador_colunas(endereco, abaplanilha):
    planilha_atual = pd.read_excel(endereco, sheet_name=abaplanilha)
    valor = planilha_atual.shape[1]
    return valor + 1


def atribuir_equipamento(consulta_icj, dict_icj_equip):
    equipamento_extraido = dict_icj_equip[consulta_icj]
    return equipamento_extraido


def agregar_dados():
    arquivo_excel = r'C:\Users\(chave)\Desktop\Python\projetos\Projeto_planilha_Guia_Medição.xlsx'
    aba = pd.read_excel(arquivo_excel, sheet_name='Base Dados')
    resumo = pd.DataFrame(aba.groupby(['Embarcação', 'Regional1', 'Equipamento', 'Regional', 'Tipo'])['Dias'].sum())
    destino = pd.ExcelWriter('Projeto_planilha_Guia_Medição_2020_1.xlsx')
    resumo.to_excel(destino, 'Base Dados', index=True)
    destino.save()


def relacionar_porte():
    wb = load_workbook(filename='Projeto_planilha_Guia_Medição.xlsx')
    aba_porte = wb['Porte']
    dict_porte = {}
    numero_linha = aba_porte.max_row
    for n in range(2, numero_linha + 1):
        tipo = aba_porte.cell(row=n, column=1).value
        dict_porte[tipo] = aba_porte.cell(row=n, column=2).value
        dict_porte.copy()
    return dict_porte


def configurar_celula(aba):
    coluna_zeros = [8, 9, 11, 14]
    coluna_percentual = [10, 13]
    numero_linha = aba.max_row
    for linha in range(2, numero_linha + 1):
        for n in coluna_zeros:
            aba.cell(row=linha, column=n).value = 0
            dias = aba.cell(row=linha, column=n)
            dias.number_format = '0.000'

    for linha1 in range(2, numero_linha + 1):
        for t in coluna_percentual:
            percentual = aba.cell(row=linha1, column=t)
            percentual.number_format = '0.000%'


def mostrar_desempenho(numero, processo):
    lista_realizacoes = ['Dados sendo extraídos e consolidados.\nProcessando...........\n',
                         'Consolidação da frota marítima realizada com sucesso.',
                         '\nInoperâncias extraídas e distribuidas com sucesso.',
                         '\nPlanilha recalculada com sucesso.',
                         '\nElaboração da versão final da planilha guia foi realizada com sucesso.']
    tempo = 0.5
    for i in range(1, numero + 1):
        time.sleep(tempo)
        print('Etapa {}/{} de análise de dados sendo realizada.\nProcessando...........\n'.format(i, numero))
        tempo = tempo + 0.5
    time.sleep(2)
    print(lista_realizacoes[processo])


def analisar_datas(data_inoperancia, data_gopi_inicio, data_gopi_fim):
    dia_ano_inope = int(data_inoperancia.strftime('%j'))
    dia_ano1 = int(data_gopi_inicio.strftime('%j'))
    dia_ano2 = int(data_gopi_fim.strftime('%j'))
    if dia_ano1 < dia_ano_inope <= dia_ano2:
        return 'sim'
    else:
        return 'não'


def agrupar_inoperancias():
    lista_inoperante = openr('Projeto_planilha_Guia_Medição.xlsx', 'Inoperância')
    aba_gopi = lista_inoperante[0]['GOPI']
    numero_linha_inope = lista_inoperante[1].max_row
    numero_linha_gopi = aba_gopi.max_row

    for linha_gopi in range(2, numero_linha_gopi + 1):
        aba_gopi.cell(row=linha_gopi, column=22).value = 0

    for linha in range(2, numero_linha_inope + 1):
        equip_inope = lista_inoperante[1].cell(row=linha, column=1).value
        data_inope = lista_inoperante[1].cell(row=linha, column=7).value
        valor_inop = lista_inoperante[1].cell(row=linha, column=13).value
        for linha2 in range(2, numero_linha_gopi + 1):
            equip_gopi = aba_gopi.cell(row=linha2, column=2).value
            if equip_inope == equip_gopi:
                data1 = aba_gopi.cell(row=linha2, column=7).value
                data2 = aba_gopi.cell(row=linha2, column=8).value
                resultado = analisar_datas(data_inope, data1, data2)
                if resultado == 'sim':
                    valor_existente = aba_gopi.cell(row=linha2, column=22).value
                    aba_gopi.cell(row=linha2, column=22).value = valor_existente + valor_inop

    for linha_gopi1 in range(2, numero_linha_gopi + 1):
        valor_celula = aba_gopi.cell(row=linha_gopi1, column=22)
        valor_celula.number_format = '0.000'
        valor_celula1 = aba_gopi.cell(row=linha_gopi1, column=22).value
        aba_gopi.cell(row=linha_gopi1, column=22).value = valor_celula1 / 24

    closer('Dados_inoperancia.xlsx', lista_inoperante[0])


def localizar_linha(centro_custo, regional, aba, inoperancia):
    linha_previa = aba.max_row
    for linha in range(2, linha_previa + 1):
        cc1 = aba.cell(row=linha, column=3).value
        if cc1 == centro_custo:
            regional_1 = aba.cell(row=linha, column=4).value
            if regional_1 == regional:
                aba.cell(row=linha, column=8).value = inoperancia


def realocar_inoperancias():
    z = openr('Projeto_planilha_Guia_Medição.xlsx', 'Previa')
    k = openr('Dados_inoperancia.xlsx', 'GOPI')
    linha_gopi = k[1].max_row
    for linha in range(2, linha_gopi + 1):
        valor_inope = k[1].cell(row=linha, column=22).value
        if valor_inope != 0:
            cc = k[1].cell(row=linha, column=2).value
            regional_principal = k[1].cell(row=linha, column=4).value
            localizar_linha(cc, regional_principal, z[1], valor_inope)
    closer('Projeto_planilha_Guia_Medição.xlsx', z[0])


def ajustar_cabecalho():
    j = openr('Projeto_planilha_Guia_Medição.xlsx', 'Previa')
    cabecalho = ['Embarcação', 'Regional1', 'Equipamento', 'Regional', 'Tipo', 'Dias',
                 'Porte', 'Indisp', 'Medir', 'PRL Petro', 'Medir Petro', 'Centro de Custo',
                 'PRL PBLOG', 'Medir PBLOG', 'Objeto de custo', 'Liberada', 'Observações']

    for v1 in range(0, len(cabecalho)):
        j[1].cell(row=1, column=v1 + 1).value = cabecalho[v1]

    closer('Projeto_planilha_Guia_Medição.xlsx', j[0])


def calcular():
    z = openr('Projeto_planilha_Guia_Medição.xlsx', 'Previa')
    linha_preenchida = z[1].max_row
    for linha in range(2, linha_preenchida + 1):
        dias = z[1].cell(row=linha, column=6).value
        indisp = z[1].cell(row=linha, column=8).value
        z[1].cell(row=linha, column=9).value = dias - indisp
        dias_medir = z[1].cell(row=linha, column=9).value
        'calculo PB1'
        prl = z[1].cell(row=linha, column=10).value
        z[1].cell(row=linha, column=11).value = dias_medir * prl
        'calculo PB2'
        prl_pblog = z[1].cell(row=linha, column=13).value
        z[1].cell(row=linha, column=14).value = dias_medir * prl_pblog
    closer('Projeto_planilha_Guia_Medição.xlsx', z[0])


def ajustar_colunas(arquivo, aba):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    wb = excel.Workbooks.Open(arquivo)
    ws = wb.Worksheets(aba)
    ws.Columns.AutoFit()
    wb.Save()
    excel.Application.Quit()


def transferir_dados():
    numero_mes = int(input('\nQual é o numero do mês da planilha guia '
                           'de medição? (ex:1-Janeiro, 2-Fevereiro, 3-Março, ...)'))

    perg2 = int(input('Qual é a revisão da planilha guia de medição? (ex:1, 2, 3,...)'))
    print('Fatorando valores a serem medidos e preparando a versão '
          'final.\npor favor aguarde!!!.\nProcessando................')
    z = openr('Projeto_planilha_Guia_Medição.xlsx', 'Previa')
    aba_medicao = z[0]['Medição']
    lista_medicao = []
    lista_dados = []
    numero_linhas = z[1].max_row
    for i in range(1, numero_linhas + 1):
        equip = z[1].cell(row=i, column=3).value
        lista_dados.append(equip)
        barco = z[1].cell(row=i, column=1).value
        lista_dados.append(barco)
        gerente = 'Gerente'
        lista_dados.append(gerente)
        fiscal = 'Fiscal'
        lista_dados.append(fiscal)
        dias = z[1].cell(row=i, column=6).value
        lista_dados.append(dias)
        indisp = z[1].cell(row=i, column=8).value
        lista_dados.append(indisp)
        medir = z[1].cell(row=i, column=9).value
        lista_dados.append(medir)
        prl = z[1].cell(row=i, column=10).value
        lista_dados.append(prl)
        dias_petro = z[1].cell(row=i, column=11).value
        lista_dados.append(dias_petro)
        centro = z[1].cell(row=i, column=12).value
        lista_dados.append(centro)
        prl_pblog = z[1].cell(row=i, column=13).value
        lista_dados.append(prl_pblog)
        dias_pblog = z[1].cell(row=i, column=14).value
        lista_dados.append(dias_pblog)
        centro1 = z[1].cell(row=i, column=15).value
        lista_dados.append(centro1)
        liberada = z[1].cell(row=i, column=16).value
        lista_dados.append(liberada)
        obs = z[1].cell(row=i, column=17).value
        lista_dados.append(obs)
        lista_medicao.append(lista_dados)
        lista_dados = []

    t = 2
    for dados in lista_medicao:
        for coluna in range(0, 15):
            if coluna == 1 or coluna == 12:
                if t > 2:
                    aba_medicao.cell(row=t, column=coluna + 1).value = dados[coluna].upper()
                else:
                    aba_medicao.cell(row=t, column=coluna + 1).value = dados[coluna]
            else:
                aba_medicao.cell(row=t, column=coluna + 1).value = dados[coluna]
        t = t + 1

    numero_linhas_medicao = aba_medicao.max_row
    for t1 in range(3, numero_linhas_medicao + 1):
        dias = aba_medicao.cell(row=t1, column=5)
        dias.number_format = '0.000'
        indisp = aba_medicao.cell(row=t1, column=6)
        indisp.number_format = '0.000'
        medir = aba_medicao.cell(row=t1, column=7)
        medir.number_format = '0.000'
        rateio = aba_medicao.cell(row=t1, column=8)
        rateio.number_format = '0.000%'
        petro_dias = aba_medicao.cell(row=t1, column=9)
        petro_dias.number_format = '0.000'
        petro_dias.font = Font(bold=True)
        rateio1 = aba_medicao.cell(row=t1, column=11)
        rateio1.number_format = '0.000%'
        pblog_dias = aba_medicao.cell(row=t1, column=12)
        pblog_dias.number_format = '0.000'
        pblog_dias.font = Font(bold=True)

    centro_custo = Planguia_modulo_2.CentroCusto()
    resultado = centro_custo.extrair_info_contrato()
    numero_linhas = aba_medicao.max_row
    for linha in range(3, numero_linhas + 1):
        numero_equip = aba_medicao.cell(row=linha, column=1).value
        aba_medicao.cell(row=linha, column=3).value = resultado[numero_equip][1]
        aba_medicao.cell(row=linha, column=4).value = resultado[numero_equip][2]

    parametro_mes = definir_mes(numero_mes)
    aba_medicao.cell(row=1, column=1).value = 'Competência:'
    aba_medicao.cell(row=1, column=2).value = parametro_mes[0]
    aba_medicao.cell(row=1, column=3).value = 'Período de medição:'
    aba_medicao.cell(row=1, column=4).value = parametro_mes[2]
    aba_medicao.cell(row=1, column=6).value = 'Revisão: '
    aba_medicao.cell(row=1, column=7).value = '0' + str(perg2)
    aba_medicao.cell(row=1, column=8).value = 'PB1'
    aba_medicao.cell(row=1, column=11).value = 'PB2'
    aba_medicao.cell(row=1, column=16).value = numero_mes

    formatar(aba_medicao)
    formatar_bordas(aba_medicao)
    formatar_ajustes(aba_medicao)
    closer('Projeto_planilha_Guia_Medição.xlsx', z[0])


def definir_mes(numero):
    t = time.localtime()
    resposta = []
    if numero == 1:
        periodo = '26/12/' + str(t[0] - 1) + ' a 25/' + str(numero) + '/' + str(t[0])
    else:
        periodo = '26/' + str(numero - 1) + '/' + str(t[0]) + ' a 25/' + str(numero) + '/' + str(t[0])
    meses_ano = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio',
                 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    mes = meses_ano[numero - 1]
    frase1 = mes + ' de ' + str(t[0])
    resposta.append(frase1)
    frase2 = '(Protocolo de envio - Planilha guia encaminhada no dia ' + str(t[2]) + '/' + str(t[1]) + '/' + str(
        t[0]) + ' as ' + str(t[3]) + ':' + str(t[4]) + ':' + str(t[5]) + ')'
    resposta.append(frase2)
    frase3 = periodo
    resposta.append(frase3)
    return resposta


def formatar(aba):
    numero_colunas = aba.max_column
    numero_linha = aba.max_row
    for v1 in range(1, numero_colunas + 1):
        for v2 in range(1, numero_linha + 1):
            aba.cell(row=v2, column=v1).alignment = Alignment(horizontal='center')

    for n in range(1, 8):
        cores_cinza = PatternFill(fill_type='solid', start_color='d6d6c2', end_color='d6d6c2')
        aba.cell(row=1, column=n).fill = cores_cinza
        aba.cell(row=1, column=n).font = Font(bold=True)
        cores_branca = PatternFill(fill_type='solid', start_color='004080', end_color='004080')
        aba.cell(row=2, column=n).fill = cores_branca
        aba.cell(row=2, column=n).font = Font(bold=True, color="ffffff")

    for n1 in range(8, 11):
        cores_azul = PatternFill(fill_type='solid', start_color='80dfff', end_color='80dfff')
        aba.cell(row=1, column=n1).fill = cores_azul
        aba.cell(row=1, column=n1).font = Font(bold=True)
        aba.cell(row=2, column=n1).fill = cores_azul

    for n2 in range(11, 14):
        cores_azulc = PatternFill(fill_type='solid', start_color='ccf2ff', end_color='ccf2ff')
        aba.cell(row=1, column=n2).fill = cores_azulc
        aba.cell(row=1, column=n2).font = Font(bold=True)
        aba.cell(row=2, column=n2).fill = cores_azulc

    for n3 in range(8, 16):
        aba.cell(row=2, column=n3).font = Font(bold=True)

    cores3 = PatternFill(fill_type='solid', start_color='ff0000', end_color='ff0000')
    aba.cell(row=2, column=14).fill = cores3
    aba.merge_cells('h1:j1')
    aba.merge_cells('k1:m1')


def formatar_bordas(ws2):
    contar_linha = ws2.max_row
    contar_colunas = ws2.max_column
    for n8 in range(2, contar_linha + 1):
        thin_lateral = Border(left=None, right=Side(style='thin'), top=None, bottom=None)
        border_1 = Border(left=None, right=Side(style='thin'), top=Side(style='thin'), bottom=None)
        ws2.cell(row=n8, column=7).border = thin_lateral
        ws2.cell(row=1, column=7).border = border_1
        ws2.cell(row=n8, column=10).border = thin_lateral
        ws2.cell(row=1, column=10).border = border_1
        ws2.cell(row=n8, column=13).border = thin_lateral
        ws2.cell(row=1, column=13).border = border_1
        ws2.cell(row=n8, column=14).border = thin_lateral
        ws2.cell(row=n8, column=15).border = thin_lateral

    for n7 in range(1, contar_colunas - 2):
        thin_topo = Border(left=None, right=None, top=Side(style='thin'), bottom=None)
        thin_bottom = Border(left=None, right=None, top=None, bottom=Side(style='thin'))
        ws2.cell(row=2, column=n7).border = thin_topo
        ws2.cell(row=contar_linha, column=n7).border = thin_bottom

    thin_lateral1 = Border(left=None, right=Side(style='thin'), top=Side(style='thin'), bottom=None)
    ws2.cell(row=2, column=7).border = thin_lateral1
    ws2.cell(row=2, column=10).border = thin_lateral1
    ws2.cell(row=2, column=13).border = thin_lateral1
    ws2.cell(row=2, column=14).border = thin_lateral1
    ws2.cell(row=2, column=15).border = thin_lateral1

    thin_bottom1 = Border(left=None, right=Side(style='thin'), top=None, bottom=Side(style='thin'))
    ws2.cell(row=contar_linha, column=7).border = thin_bottom1
    ws2.cell(row=contar_linha, column=10).border = thin_bottom1
    ws2.cell(row=contar_linha, column=13).border = thin_bottom1
    ws2.cell(row=contar_linha, column=14).border = thin_bottom1
    ws2.cell(row=contar_linha, column=15).border = thin_bottom1


def formatar_ajustes(aba):
    contar_numero_linha = aba.max_row
    for n in range(3, contar_numero_linha+1):
        prl_petro = aba.cell(row=n, column=8).value
        if prl_petro == 0:
            aba.cell(row=n, column=10).value = '-'
            aba.cell(row=n, column=10).alignment = Alignment(horizontal='center')
        elif prl_petro == 1:
            aba.cell(row=n, column=13).value = '-'
            aba.cell(row=n, column=13).alignment = Alignment(horizontal='center')

    for t in range(3, contar_numero_linha+1):
        embarc_liberada = aba.cell(row=t, column=14).value
        if embarc_liberada is None:
            aba.cell(row=t, column=14).value = 'Sim'
            aba.cell(row=t, column=14).alignment = Alignment(horizontal='center')
    for w in range(3, contar_numero_linha+1):
        obsercacao = aba.cell(row=w, column=15).value
        if obsercacao is None:
            aba.cell(row=w, column=15).value = '-'
            aba.cell(row=w, column=15).alignment = Alignment(horizontal='center')

    branco = PatternFill(fill_type='solid', start_color='ffffff', end_color='ffffff')
    contar_numero_linha1 = aba.max_row
    contar_numero_coluna1 = aba.max_column
    for linha in range(3, contar_numero_linha1+1):
        for coluna in range(1, contar_numero_coluna1):
            aba.cell(row=linha, column=coluna).fill = branco

    contar_linha2 = aba.max_row
    for n9 in range(3, contar_linha2 + 1):
        status_medicao = aba.cell(row=n9, column=14).value
        if status_medicao == 'Não':
            aba.cell(row=n9, column=15).value = 'Embarcação possui dias sob o status de suspensão contratual'
            for n10 in range(1, 16):
                cores5 = PatternFill(fill_type='solid', start_color='ff9999', end_color='ff9999')
                aba.cell(row=n9, column=n10).fill = cores5


def saudar():
    t = time.localtime()
    z = t[3]
    if z < 12:
        a = 'Prezados Gerentes e Fiscais de contrato, Bom dia!'
        return a
    elif z >= 18:
        b = 'Prezados Gerentes e Fiscais de contrato, Boa noite!'
        return b
    else:
        c = 'Prezados Gerentes e Fiscais de contrato, Boa tarde!'
        return c


def converter_moeda(valor):
    valor_final = format_currency(valor, 'USD', '¤¤ #,##0.00', locale='es_ES')
    return valor_final


def importar_cambio():
    request = requests.get('https://economia.awesomeapi.com.br/json/all')
    cotacao = json.loads(request.text)
    valor_compra = (cotacao['USD']['high'])
    return float(valor_compra)


def deletar(arquivo, aba):
    t = openr(arquivo, aba)
    t[1].delete_cols(1, 17)
    closer(arquivo, t[0])
