from tkinter import messagebox
import pandas as pd
from openpyxl import load_workbook


class PlanInitial:
    def __init__(self, pathprinc, pathdest, dataprl, prt, contr):
        self.pfile = pathprinc
        self.destination = pathdest
        self.prl = dataprl
        self.port = prt
        self.contract = contr

    def agregdatamedic(self):
        """
        Prepara um dataframe com os dados relevantes, retiradas do arquivo do GOPI.
        """
        df = pd.read_excel(self.pfile, sheet_name='GOPI', skiprows=2)
        df = df[['ICJ', 'Embarcação', 'Classe', 'APONTAMENTO\n(PARA GOPI/PAD)',
                 'MEDIÇÃO\n (PARA GCI/CMAR)', 'Dias de Barcos relativos']]
        df['Dias de Barcos relativos'] = df['Dias de Barcos relativos'].round(3)
        df.sort_values(by=['Embarcação'], ascending=True, inplace=True)
        df.rename(columns={'APONTAMENTO\n(PARA GOPI/PAD)': 'Regional', 'MEDIÇÃO\n (PARA GCI/CMAR)': 'Regional CMAR',
                           'Dias de Barcos relativos': 'Dias Medir'}, inplace=True)
        df.to_excel(self.destination, index=False)

    def validcontract(self):
        """
        Associa numero de equipamento buscando o numero de ICJ de cada embarcação.
        """
        try:
            wb = load_workbook(self.destination)
            ws = wb['Sheet1']
            ws.cell(row=1, column=7).value = 'Equipamento'
            contador = ws.max_row
            for linha in range(2, contador + 1):
                icj = ws.cell(row=linha, column=1).value
                ws.cell(row=linha, column=7).value = self.contract[icj][0]
            wb.save(self.destination)
        except Exception as err:

            warning = f'Erro: Validação rejeitada!!!\nICJ {err} inválido. Por favor verifique ' \
                      f'se este ICJ encontra-se previamente cadastrado!!!.'
            handleerror(warning)

    def validport(self):
        """
        Associa o porte da embarcação pelo tipo.
        """
        embarc = None
        try:
            wb = load_workbook(self.destination)
            ws = wb['Sheet1']
            ws.cell(row=1, column=8).value = 'Porte'
            contador = ws.max_row
            for linha in range(2, contador + 1):
                embarc = ws.cell(row=linha, column=2).value
                portet = ws.cell(row=linha, column=3).value
                ws.cell(row=linha, column=8).value = self.port[portet]
            wb.save(self.destination)
        except Exception as err:
            warning = f'Erro: Validação Rejeitada!!!.\nNão existe porte de embarcação relacionado a {err}, ' \
                      f'atribuído a embarcação {embarc}.'
            handleerror(warning)

    def alocatedataship(self):
        """
        Função que realiza a alocação de dados de cada embarcação
        """
        linha = None
        try:
            wb = load_workbook(self.destination)
            ws = wb['Sheet1']
            listreg = self.factory3()
            contador = ws.max_row
            for linha in range(2, contador + 1):
                baseicj = ws.cell(row=linha, column=1).value
                regional = ws.cell(row=linha, column=5).value
                pte = ws.cell(row=linha, column=8).value
                if regional == 'B. Campos ES' or regional == 'B. Santos' or regional == 'P. Búzios':
                    self.alocatereg(ws, linha, baseicj, regional, pte)
                else:
                    self.alocateregothers(ws, linha, baseicj, regional, listreg)
            wb.save(self.destination)
        except Exception as err:
            warning = f'Erro:\nA regional {err}, informado na linha {linha} não possui critério de rateio.'
            handleerror(warning)

            # exit()

    def alocatereg(self, aba, line, baseicj, regional, pte):
        """
        Aloca os dados das embarcaçõe das três principais regionais.
        """
        if regional == 'P. Búzios':
            regio = 'B. Santos'
        else:
            regio = regional
        listre = self.factory2(regio, pte)
        ind = self.prl[listre]
        aba.cell(row=line, column=11).value = ind['PRL Petro']
        aba.cell(row=line, column=13).value = ind['Objeto de custo']
        aba.cell(row=line, column=14).value = ind['PRL PBLog']
        aba.cell(row=line, column=17).value = self.analisecontract(baseicj)[2]
        aba.cell(row=line, column=18).value = self.analisecontract(baseicj)[0]
        aba.cell(row=line, column=19).value = self.analisecontract(baseicj)[1]
        if self.analisecontract(baseicj)[2] == 'Sim':
            aba.cell(row=line, column=16).value = ind['Pólo (Tem cessão)']
        else:
            aba.cell(row=line, column=16).value = ind['Objeto de custo (Não tem cessão)']

    def alocateregothers(self, aba, line, baseicj, regional, listregional):
        """
        Aloca os dados das embarcações das regionais secundárias.
        """
        index = listregional.index(regional)
        ind2 = self.prl[index]
        aba.cell(row=line, column=11).value = ind2['PRL Petro']
        aba.cell(row=line, column=13).value = ind2['Objeto de custo']
        aba.cell(row=line, column=14).value = ind2['PRL PBLog']
        aba.cell(row=line, column=17).value = self.analisecontract(baseicj)[2]
        aba.cell(row=line, column=18).value = self.analisecontract(baseicj)[0]
        aba.cell(row=line, column=19).value = self.analisecontract(baseicj)[1]
        if self.analisecontract(baseicj)[2] == 'Sim':
            aba.cell(row=line, column=16).value = ind2['Pólo (Tem cessão)']
        else:
            aba.cell(row=line, column=16).value = ind2['Objeto de custo (Não tem cessão)']

    @staticmethod
    def factory2(reg, pt):
        """
        Função que fornece a chave do dicionário com os dados de medição.
        """
        height = ['EPP', 'EMP', 'EGP']
        if reg == 'B. Campos ES':
            return height.index(pt)
        elif reg == 'B. Santos':
            return height.index(pt) + 3

    def factory3(self):
        """
        Função que fornece umaa lista com as regionais e seus dados.
        """
        wb = load_workbook(self.pfile)
        ws = wb['PRL']
        contador = ws.max_row
        listregional = []
        for linha in range(3, contador + 1):
            regional = ws.cell(row=linha, column=2).value
            listregional.append(regional)
        return listregional

    def analisecontract(self, base):
        """
        Função que fornece o nome do gerente e fiscal de contrato de cada embarcação usando o ICJ.
        """
        infocontract = self.contract[base]
        return infocontract[2], infocontract[3], infocontract[4]

    def agregacion(self):
        aba = pd.read_excel(self.destination, sheet_name='Sheet1')
        resumo = pd.DataFrame(aba.groupby(['ICJ', 'Embarcação', 'Classe', 'Regional',
                                           'Regional CMAR'])['Dias Medir'].sum()).round(4)
        destino = pd.ExcelWriter(self.destination)
        resumo.to_excel(destino, 'Sheet1', index=True)
        destino.save()
        df = pd.read_excel(self.destination, sheet_name='Sheet1')
        df = df[['ICJ', 'Embarcação', 'Classe', 'Regional', 'Regional CMAR', 'Dias Medir']]
        df.to_excel(self.destination, index=False)

    def ajustar_celulas(self):
        wb = load_workbook(self.destination)
        ws = wb['Sheet1']
        contar_linha = ws.max_row
        for n in range(2, contar_linha + 1):
            for t in range(1, 4):
                atual = ws.cell(row=n, column=t).value
                if atual is None:
                    anterior = ws.cell(row=n - 1, column=t).value
                    ws.cell(row=n, column=t).value = anterior
        wb.save(self.destination)

    def formatation(self):
        df = pd.read_excel(self.destination, sheet_name='Sheet1')
        df.rename(columns={'Unnamed: 6': 'Equipamento', 'Unnamed: 7': 'Porte', 'Unnamed: 8': 'Indisp',
                           'Unnamed: 9': 'Medir', 'Unnamed: 10': 'PRL Petro', 'Unnamed: 11': 'Medir Petro',
                           'Unnamed: 12': 'Centro de Custo', 'Unnamed: 13': 'PRL PBLOG', 'Unnamed: 14': 'Medir PBLOG',
                           'Unnamed: 15': 'Objeto de Custo', 'Unnamed: 16': 'Cessão', 'Unnamed: 17': 'Gerente',
                           'Unnamed: 18': 'Fiscal'}, inplace=True)
        for i, row in df.iterrows():
            if row['Regional CMAR'] == 'Suspenso':
                df.loc[i, 'Autorizado'] = 'Não'
            else:
                df.loc[i, 'Autorizado'] = 'Sim'
        df.sort_values(by=['Embarcação'], ascending=True, inplace=True)
        df['Observações'] = '-'
        df.to_excel(self.destination, index=False)


def handleerror(err):
    messagebox.showerror(title='Mensagem de erro', message=err)
