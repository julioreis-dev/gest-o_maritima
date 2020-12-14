import win32com.client
from openpyxl import load_workbook
from modulo_calculo import CalcPlanGui
import time


class Email(CalcPlanGui):
    def __init__(self, pathorigin, pathdest):
        super().__init__(pathorigin, pathdest)


    def planinformation(self):
        t = time.localtime()
        resposta = []
        month = input(
            '\nQual é o numero do mês da planilha guia de medição? (ex:1-Janeiro, 2-Fevereiro, 3-Março, ...)?')
        review = input('Qual é a revisão da planilha guia de medição? (ex:1, 2, 3,...)?')
        if review.isdigit():
            review = int(review)
        else:
            print('Erro ao informar a revisão.')

        if month.isdigit():
            month = int(month)
            if month == 1:
                period = '26/12/' + str(t[0] - 1) + ' a 25/' + str(month) + '/' + str(t[0])
            else:
                period = '26/' + str(month - 1) + '/' + str(t[0]) + ' a 25/' + str(month) + '/' + str(t[0])
            meses_ano = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio',
                         'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            mes = meses_ano[month - 1]
            frase1 = mes + ' de ' + str(t[0])
            resposta.append(frase1)
            frase2 = '(Protocolo de envio - Planilha guia encaminhada no dia ' + str(t[2]) + '/' + str(t[1]) + '/' + \
                     str(t[0]) + ' as ' + str(t[3]) + ':' + str(t[4]) + ':' + str(t[5]) + ')'
            frase3 = period
            resposta.append(frase3)
            resposta.append('0' + str(review))
            resposta.append(frase2)
        return resposta

    @staticmethod
    def filesname(namesdata):
        filename = namesdata[0] + '- Planilha Guia_R' + namesdata[2] + '.xlsx'
        return filename


    def clearcell(self):
        wb = load_workbook(self.pfile1)
        ws = wb['Medição']
        contador = ws.max_row
        for col in range(1, 22):
            for linha in range(3, contador+1):
                ws.cell(row=linha, column=col).value = ''
        wb.save(self.pfile1)

    def preparedest(self):
        wb = load_workbook(self.pfile3)
        ws = wb['Chaves']
        destiny = []
        copy = []
        cont = ws.max_row
        for w in range(2, cont + 1):
            status = ws.cell(row=w, column=6).value
            if status == 'Sim':
                destiny.append(ws.cell(row=w, column=1).value)
            elif status == 'Copy':
                copy.append(ws.cell(row=w, column=1).value)
            else:
                pass
        finaldestiny = '; '.join(destiny)
        finalcopy = '; '.join(copy)
        return finaldestiny, finalcopy

    def planformat(self, listdata):
        wb = load_workbook(self.pfile3)
        ws = wb['Medição']
        fields = [2, 5, 8]
        for i, n in enumerate(fields):
            ws.cell(row=1, column=n).value = listdata[i]
        wb.save(self.pfile3)

    def sendemail(self, rev, dest, content, anx):
        try:
            anexo = anx
            o = win32com.client.Dispatch("Outlook.Application")
            msg = o.CreateItem(0)
            msg.To = dest[0]
            msg.CC = dest[1]
            msg.BCC = ''
            msg.Subject = content[1]
            msg.Body = content[0] + '\n' + content[2] + '\n' + rev[3]
            msg.Attachments.Add(anexo)
            index = anexo.rfind('/')
            extensao = anexo[index + 1:]
            msg.Send()
            return 'Email enviado com sucesso!'
        except Exception as err:
            return f'Erro: {err}\nDurante o processo de elaboração e envio do email um erro foi reportado!!!'

    @staticmethod
    def saudar():
        t = time.localtime()
        z = t[3]
        if z < 12:
            a = 'Prezados Gerentes e Fiscais de contrato, Bom dia!'
            return a
        elif z >= 18:
            b = 'Prezados Gerentes e Fiscais de contrato, Boa noite!'
            return b
        else:
            c = 'Prezados Gerentes e Fiscais de contrato, Boa tarde!'
            return c

    def emailcontent(self, rev):
        corpo_email = self.saudar() + '\n\nSegue anexo a planilha guia de medição (Revisão ' + str(
            rev[2]) + ').\nQualquer ' \
                      'informação a acrescentar em alguma embarcação da frota, ' \
                      'que não constar na planilha, favor avisar para ajuste.' \
                      '\n\nATENÇÃO: PARA EMBARCAÇÕES QUE NÃO FORAM LIBERADAS ' \
                      'PARA MEDIÇÃO VER NA COLUNA (N - "Embarcação Liberada") DA ABA "MEDIÇÃO".' \
                      ' TODAS AS ALTERAÇÕES REALIZADAS ' \
                      'NA PLANILHA GUIA SÃO REGISTRADAS NA ABA "REVISÃO".' \
                      '\n\nAs embarcações "Não Liberadas" estão sendo analisadas ' \
                      'pela equipe de coordenação e controle das informações ' \
                      'referente a frota sob a gestão do CMAR, e em breve essas ' \
                      'embarcações poderão sofrer alteração quanto ao seu status. ' \
                      'Caso isso ocorra, uma nova revisão será emitida ' \
                      'com a finalidade de ajustar a planilha guia.'

        subject = 'CMAR - Planilha Guia de Medição - ' + rev[0] + ' - REVISÃO ' + rev[2]
        body = '\nHistórico de ajustes realizados - Revisão ' + rev[2] + ':\n' \
               + self.listar_revisao() + '\nAtenciosamente,\nEquipe de Gerenciamento Marítimo\nLOEP/LOFF/GCI/CMAR\n'
        return corpo_email, subject, body

    def listar_revisao(self):
        wb = load_workbook(self.pfile3)
        ws = wb['Revisão']
        lista = []
        numero_linhas = ws.max_row
        for n in range(2, numero_linhas + 1):
            status = ws.cell(row=n, column=3).value
            if status == 'ok':
                alter = ws.cell(row=n, column=2).value
                lista.append(alter)
        sentenca = ''
        for w in range(0, len(lista)):
            valor = lista[w]
            sentenca = sentenca + str(valor) + '\n'
        return sentenca
