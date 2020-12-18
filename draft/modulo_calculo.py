import pandas as pd
from openpyxl import load_workbook
import openpyxl.styles as styles


class CalcPlanGui:
    def __init__(self, pathorigin, pathdest):
        self.pfile1 = pathorigin
        self.pfile2 = pathdest
        self.pfile3 = r'C:\Users\ay4m\Desktop\planguia\arquivo_editado.xlsx'

    def calcdata(self):
        df = pd.read_excel(self.pfile2, sheet_name='Sheet1')
        df = df.fillna(0.00)
        df['Medir'] = (df['Dias Medir'] - df['Indisp']).round(3)
        df['Medir Petro'] = (df['Medir'] * df['PRL Petro']).round(3)
        df['Medir PBLOG'] = (df['Medir'] * df['PRL PBLOG']).round(3)
        df.to_excel(self.pfile2, index=False)
        # self.sheetconfiguration()
        # self.finalversion()
        # self.sendproduct()

    def sheetconfiguration(self):
        wb = load_workbook(self.pfile2)
        ws = wb['Sheet1']
        coluna_zeros = [9, 10, 12, 15]
        coluna_percentual = [11, 14]
        numero_linha = ws.max_row
        for linha in range(2, numero_linha + 1):
            for n in coluna_zeros:
                dias = ws.cell(row=linha, column=n)
                dias.number_format = '0.000'
        for linha1 in range(2, numero_linha + 1):
            for t in coluna_percentual:
                percentual = ws.cell(row=linha1, column=t)
                percentual.number_format = '0.000%'
        wb.save(self.pfile2)

    def finalversion(self):
        df = pd.read_excel(self.pfile2, sheet_name='Sheet1')
        df1 = df[['Equipamento', 'Embarcação', 'Regional CMAR', 'Gerente', 'Fiscal', 'Dias Medir', 'Indisp', 'Medir',
                  'PRL Petro',
                  'Medir Petro', 'Centro de Custo', 'PRL PBLOG', 'Medir PBLOG', 'Objeto de Custo', 'Autorizado',
                  'Observações']]

        # carrego o Excel com o template pré-formatado 'template.xlsx'
        book = load_workbook(self.pfile1)

        # defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
        writer = pd.ExcelWriter(self.pfile3, engine='openpyxl')

        # incluo a formatação no writer
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

        # Escrevo com o .to_excel() do pandas
        df.to_excel(writer, 'Previa', index=False)

        # para escrever só os valores em um lugar específico:
        df1.to_excel(writer, 'Medição', startrow=2, startcol=0, header=False, index=False)

        writer.save()

    def sendproduct(self):
        wb = load_workbook(self.pfile3)
        ws = wb['Medição']
        numero_linha = ws.max_row
        for linha2 in range(2, numero_linha + 1):
            susp = ws.cell(row=linha2, column=3).value
            if susp == 'Suspenso':
                for col in range(1, 16):
                    ws.cell(row=linha2, column=col).font = styles.Font(bold=True, color="FF0000")
        wb.save(self.pfile3)
