import win32com.client
from openpyxl import load_workbook
import Planguia_funcoes


class EnviarEmail:
    def __init__(self, arquivo, aba, mes_ano, revisao):
        self.arquivo = arquivo
        self.aba = aba
        self.mes = mes_ano
        self.revisao = revisao

    def preparar_email(self):
        aba_porte = self.arquivo.get_sheet_by_name('Porte')
        self.arquivo.remove_sheet(aba_porte)
        aba_taxa = self.arquivo.get_sheet_by_name('Taxa Diária')
        self.arquivo.remove_sheet(aba_taxa)
        aba_previa = self.arquivo.get_sheet_by_name('Previa')
        self.arquivo.remove_sheet(aba_previa)
        aba_info_pblog = self.arquivo.get_sheet_by_name('Info Contrato - Pblog')
        self.arquivo.remove_sheet(aba_info_pblog)
        aba_chaves = self.arquivo.get_sheet_by_name('Chaves')
        self.arquivo.remove_sheet(aba_chaves)
        dados_rev = self.aba.cell(row=1, column=7).value
        mes_medicao = self.aba.cell(row=1, column=2).value
        self.aba.cell(row=1, column=16).value = ''
        dest_filename = mes_medicao + ' - Planilha Guia_R' + str(dados_rev) + '.xlsx'
        self.arquivo.save(filename=dest_filename)
        self.arquivo.close()
        return dest_filename

    def listar_revisao(self):
        ws2 = self.arquivo['Revisão']
        lista = []
        numero_linhas = self.aba.max_row
        for n in range(2, numero_linhas + 1):
            status = ws2.cell(row=n, column=3).value
            if status == 'ok':
                alter = ws2.cell(row=n, column=2).value
                lista.append(alter)
        sentenca = ''
        for w in range(0, len(lista)):
            valor = lista[w]
            sentenca = sentenca + str(valor) + '\n'
        return sentenca

    def listar_destinatario(self):
        t = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição.xlsx', 'Chaves')
        contar_dest = t[1].max_row
        lista_destinatario = []
        lista_destinatario1 = []
        lista_resultado = []
        for n in range(2, contar_dest + 1):
            status = t[1].cell(row=n, column=6).value
            if status == 'Sim':
                endereco = t[1].cell(row=n, column=1).value
                lista_destinatario.append(endereco)
            elif status == 'Copy':
                endereco2 = t[1].cell(row=n, column=1).value
                lista_destinatario1.append(endereco2)
        saudacao = Planguia_funcoes.saudar()
        destinatarios = '; '.join(lista_destinatario)
        destinatarios2 = '; '.join(lista_destinatario1)
        lista_resultado.append(destinatarios)
        corpo_email = saudacao + '\n\nSegue anexo a planilha guia de medição (Revisão ' + str(
            self.revisao) + ').\nQualquer informação a acrescentar em alguma embarcação da frota, ' \
                            'que não constar na planilha, favor avisar para ajuste.' \
                            '\n\nATENÇÃO: PARA EMBARCAÇÕES QUE NÃO FORAM LIBERADAS ' \
                            'PARA MEDIÇÃO VER NA COLUNA (N - "Embarcação Liberada") DA ABA "MEDIÇÃO".' \
                            ' TODAS AS ALTERAÇÕES REALIZADAS ' \
                            'NA PLANILHA GUIA SÃO REGISTRADAS NA ABA "REVISÃO".' \
                            '\n\nAs embarcações "Não Liberadas" estão sendo analisadas ' \
                            'pela equipe de coordenação e controle das informações ' \
                            'referente a frota sob a gestão do CMAR, e em breve essas ' \
                            'embarcações poderão sofrer alteração quanto ao seu status. ' \
                            'Caso isso ocorra, uma nova revisão será emitida ' \
                            'com a finalidade de ajustar a planilha guia.'
        lista_resultado.append(corpo_email)
        lista_resultado.append(destinatarios2)
        return lista_resultado

    def encaminhar_email(self):
        try:
            anexo = r'C:/Users/ay4m/Desktop/Python/projetos/' + self.preparar_email()
            lista_destinatario = self.listar_destinatario()
            o = win32com.client.Dispatch("Outlook.Application")
            msg = o.CreateItem(0)
            msg.To = lista_destinatario[0]
            msg.CC = lista_destinatario[2]
            msg.BCC = ''
            msg.Subject = 'CMAR - Planilha Guia de Medição - ' + self.mes[0] + ' - REVISÃO ' + str(self.revisao)
            msg.Body = lista_destinatario[1] + '\nHistórico de ajustes realizados - Revisão ' + str(
                self.revisao) + ':\n' + self.listar_revisao() + '\nAtenciosamente,' \
                                                                '\nEquipe de Gerenciamento Marítimo\nLOEP/LOFF/GCI/CMAR\n' + self.mes[1]
            msg.Attachments.Add(anexo)
            msg.Send()
            return '\nEmail enviado com sucesso!'
        except:
            return 'Erro ao enviar email!!!'


def iniciar_3():
    t = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição.xlsx', 'Medição')
    numero_mes = t[1].cell(row=1, column=16).value
    mes = Planguia_funcoes.definir_mes(numero_mes)
    rev = t[1].cell(row=1, column=7).value
    email = EnviarEmail(t[0], t[1], mes, rev)
    resposta = email.encaminhar_email()
    print(resposta)
