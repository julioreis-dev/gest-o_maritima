import Planguia_funcoes
from openpyxl.styles import Font, numbers
import time

"""" Esta classe molda todos os relacionamentos na alocação de centro de custos e outras informações relevantes."""


class CentroCusto:
    def __init__(self, regional=0, porte=0, regional1=0, dict_dados=dict):
        self.regional = regional
        self.porte = porte
        self.regional1 = regional1
        self.lista_pronta = dict_dados

    def definir_centroregional(self):

        """ Retorna valores dos metodos: tratar_ols_olnf, tratar_ols_olnne, tratar_outrasregionais."""

        if self.regional == 'OLS' or self.regional == 'OLNF':
            resultado = self.tratar_ols_olnf()
            return resultado
        elif self.regional == 'OLNNE':
            resultado1 = self.tratar_olnne()
            return resultado1
        else:
            resultado2 = self.tratar_outrasregionais()
            return resultado2

    def tratar_ols_olnf(self):
        """ Retorna uma lista com os dados relevantes para a medição referentes a OLS ou OLNF."""
        infor_reg = self.lista_pronta
        informacoes = infor_reg[self.regional]
        if self.porte == 'EPP':
            return informacoes[0]
        elif self.porte == 'EMP':
            return informacoes[1]
        elif self.porte == 'EGP':
            return informacoes[2]

    def tratar_olnne(self):
        """ Retorna uma lista com os dados relevantes para a medição referente a OLNNE."""
        infor_reg = self.lista_pronta
        informacoes = infor_reg[self.regional]
        if self.regional1 == 'SEAL':
            return informacoes[0]
        elif self.regional1 == 'UO-RNCE':
            return informacoes[1]
        elif self.regional1 == 'UO-BA':
            return informacoes[2]
        elif self.regional1 == 'Área Remota':
            return informacoes[3]

    def tratar_outrasregionais(self):
        """ Retorna uma lista com os dados relevantes para a medição referente as outras regionais."""
        infor_reg = self.lista_pronta
        informacoes = infor_reg[self.regional]
        return informacoes[0]

    def organizar_regional(self):
        """ Retorna um lista de reginais (sem repetição) que possuem operação maritima."""
        a = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020.xlsx', 'PRL')
        lista_regional = []
        numero_linha = a[1].max_row
        for n in range(3, numero_linha + 1):
            reg = a[1].cell(row=n, column=2).value
            lista_regional.append(reg)
        self.lista_pronta = sorted(set(lista_regional))
        return self.lista_pronta

    def catalogar_cc(self):
        """ Retorna um dicionário.
        ex:{Regional:[[Reg1, Reg, CC, CC-PBLOG, atv, None, Obj(com cessão), Obj(sem cessão), None, PRL, PRL PBLOG]]}"""
        a = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020.xlsx', 'PRL')
        self.organizar_regional()
        dict_dadosregionais = {}
        lista = []
        lista_geral = []
        lista_dados = self.lista_pronta
        numero_linha = a[1].max_row
        for indice in range(0, len(lista_dados)):
            reg = lista_dados[indice]
            for n in range(3, numero_linha + 1):
                reg2 = a[1].cell(row=n, column=2).value
                if reg == reg2:
                    for w in range(1, 12):
                        info = a[1].cell(row=n, column=w).value
                        lista.append(info)
                    lista_geral.append(lista)
                    lista = []
            dict_dadosregionais[reg] = lista_geral
            dict_dadosregionais.copy()
            lista_geral = []
        return dict_dadosregionais

    @staticmethod
    def extrair_info_contrato():
        """ Retorna um dicionário.ex:{equipamento:[nome_embarcação, gerente, fiscal, cessão contratual]}."""
        a = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020.xlsx', 'Info Contrato')
        numero_linha = a[1].max_row
        lista_contrato = []
        dict_contrato = {}
        for n in range(2, numero_linha + 1):
            equip = a[1].cell(row=n, column=2).value
            embarcacao = a[1].cell(row=n, column=4).value
            lista_contrato.append(embarcacao)
            gerente = a[1].cell(row=n, column=6).value
            lista_contrato.append(gerente)
            fiscal = a[1].cell(row=n, column=8).value
            lista_contrato.append(fiscal)
            cessao = a[1].cell(row=n, column=12).value
            lista_contrato.append(cessao)
            dict_contrato[equip] = lista_contrato
            dict_contrato.copy()
            lista_contrato = []
        return dict_contrato

    @staticmethod
    def definir_porte():
        """ Retorna o porte relacionado a cada embarcação."""
        x = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020.xlsx', 'Previa')
        numero_linha = x[1].max_row
        dict_portes = Planguia_funcoes.relacionar_porte()
        for n in range(2, numero_linha + 1):
            tipo = x[1].cell(row=n, column=5).value
            porte = dict_portes[tipo]
            x[1].cell(row=n, column=7).value = porte
        Planguia_funcoes.closer('Projeto_planilha_Guia_Medição_2020.xlsx', x[0])


centro = CentroCusto()
centro.definir_porte()
y = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020.xlsx', 'Previa')
base_dados = y[0].get_sheet_by_name('Base Dados')
y[0].remove_sheet(base_dados)
n_linha = y[1].max_row
dados_gerais_medicao = centro.catalogar_cc()
dados_contrato = centro.extrair_info_contrato()
for linha in range(2, n_linha + 1):
    polo_equip = y[1].cell(row=linha, column=3).value
    polo_regional = y[1].cell(row=linha, column=4).value
    polo_porte = y[1].cell(row=linha, column=7).value
    polo_regional1 = y[1].cell(row=linha, column=2).value
    dados1 = CentroCusto(polo_regional, polo_porte, polo_regional1, dados_gerais_medicao)
    w1 = dados1.definir_centroregional()
    contrato_lista = dados_contrato[polo_equip]
    if contrato_lista[3] == 'Sim':
        y[1].cell(row=linha, column=10).value = w1[9]
        y[1].cell(row=linha, column=12).value = w1[2]
        y[1].cell(row=linha, column=13).value = w1[10]
        y[1].cell(row=linha, column=15).value = w1[6]
    elif contrato_lista[3] == 'Não':
        y[1].cell(row=linha, column=10).value = w1[9]
        y[1].cell(row=linha, column=12).value = w1[2]
        y[1].cell(row=linha, column=13).value = w1[10]
        y[1].cell(row=linha, column=15).value = w1[7]
Planguia_funcoes.closer('Projeto_planilha_Guia_Medição_2020.xlsx', y[0])

t = Planguia_funcoes.openr('Projeto_planilha_Guia_Medição_2020.xlsx', 'Previa')
Planguia_funcoes.configurar_celula(t[1])
Planguia_funcoes.closer('Projeto_planilha_Guia_Medição_2020.xlsx', t[0])

Planguia_funcoes.mostrar_desempenho(5, 1)
