from datetime import datetime
from openpyxl import load_workbook

origin = r'C:\Users\ay4m\Desktop\planguia\planilha_Guia_Medição.xlsx'


def analisegopi():
    lista_data = []
    wb = load_workbook(origin)
    ws = wb['GOPI']
    numero_linha = ws.max_row
    for line in range(4, numero_linha + 1):
        icj = ws.cell(row=line, column=2).value
        start = ws.cell(row=line, column=7).value
        end = ws.cell(row=line, column=8).value
        operation = ws.cell(row=line, column=9).value
        resume = (icj, start, end, operation)
        lista_data.append(resume)
    return lista_data


def analiseindisp():
    dict_indip = {}
    wb = load_workbook(origin)
    ws = wb['Inoperância']
    numero_linha = ws.max_row
    for line in range(2, numero_linha):
        lista_indisp = []
        icj = ws.cell(row=line, column=1).value
        data_start = ws.cell(row=line, column=5).value
        hr_start = ws.cell(row=line, column=6).value
        data_end = ws.cell(row=line, column=7).value
        hr_end = ws.cell(row=line, column=8).value
        duration = ws.cell(row=line, column=9).value

        data_start = data_start.strftime('%Y-%m-%d')
        hr_start = hr_start.strftime('%X')
        datetime_start = data_start + ' ' + hr_start
        datetime_start = datetime.strptime(datetime_start, '%Y-%m-%d %H:%M:%S')

        data_end = data_end.strftime('%Y-%m-%d')
        hr_end = hr_end.strftime('%X')
        datetime_end = data_end + ' ' + hr_end
        datetime_end = datetime.strptime(datetime_end, '%Y-%m-%d %H:%M:%S')

        data_indisp = (datetime_start, datetime_end, round((duration/24), 3))
        dict_indip.setdefault(icj, lista_indisp)
        consolided = dict_indip[icj]
        consolided.append(data_indisp)
        dict_indip[icj] = consolided
    return dict_indip

def separatedata(dict_indispdata, equip):
    tupladata = dict_indispdata[equip]
    for info in tupladata:
        print(info)

x = analiseindisp()
y = separatedata(x, 30096540)
# print(x)
# format = x[0][1]
# print(x[0][1])
# print(x)
