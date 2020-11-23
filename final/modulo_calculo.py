import pandas as pd
from openpyxl import load_workbook
import modulo_verfinal as final


class CalcPlanGui:
    def __init__(self, pathorigin, pathdest):
        self.pfile1 = pathorigin
        self.pfile2 = pathdest

    def calcdata(self):
        df = pd.read_excel(self.pfile2, sheet_name='Sheet1')
        df = df.fillna(0.00)
        df['Medir'] = (df['Dias Medir'] - df['Indisp']).round(3)
        df['Medir Petro'] = (df['Medir'] * df['PRL Petro']).round(3)
        df['Medir PBLOG'] = (df['Medir'] * df['PRL PBLOG']).round(3)
        df.to_excel(self.pfile2, index=False)
        self.sheetconfiguration()
        final.finalversion(self.pfile1, self.pfile2)
        final.sendproduct()

    def sheetconfiguration(self):
        wb = load_workbook(self.pfile2)
        ws = wb['Sheet1']
        coluna_zeros = [9, 10, 12, 15]
        coluna_percentual = [11, 14]
        numero_linha = ws.max_row
        for linha in range(2, numero_linha + 1):
            for n in coluna_zeros:
                dias = ws.cell(row=linha, column=n)
                dias.number_format = '0.000'
        for linha1 in range(2, numero_linha + 1):
            for t in coluna_percentual:
                percentual = ws.cell(row=linha1, column=t)
                percentual.number_format = '0.000%'
        wb.save(self.pfile2)
