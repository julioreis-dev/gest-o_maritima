import pandas as pd
from openpyxl import load_workbook
import openpyxl.styles as styles


def finalversion(file1, file2):
    df = pd.read_excel(file2, sheet_name='Sheet1')
    df1 = df[['Equipamento', 'Embarcação', 'Regional CMAR', 'Gerente', 'Fiscal', 'Dias Medir', 'Indisp', 'Medir',
              'PRL Petro',
              'Medir Petro', 'Centro de Custo', 'PRL PBLOG', 'Medir PBLOG', 'Objeto de Custo', 'Autorizado',
              'Observações']]

    # carrego o Excel com o template pré-formatado 'template.xlsx'
    book = load_workbook(file1)

    # defino o writer para escrever em um novo arquivo 'arquivo_editado.xlsx'
    writer = pd.ExcelWriter(r'C:\Users\Julio\Desktop\arquivo_editado.xlsx', engine='openpyxl')

    # incluo a formatação no writer
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Escrevo com o .to_excel() do pandas
    df.to_excel(writer, 'Previa', index=False)

    # para escrever só os valores em um lugar específico:
    df1.to_excel(writer, 'Medição', startrow=2, startcol=0, header=False, index=False)

    writer.save()


def sendproduct():
    wb = load_workbook(r'C:\Users\Julio\Desktop\arquivo_editado.xlsx')
    ws = wb['Medição']
    numero_linha = ws.max_row
    for linha2 in range(2, numero_linha + 1):
        susp = ws.cell(row=linha2, column=3).value
        if susp == 'Suspenso':
            for col in range(1, 15):
                ws.cell(row=linha2, column=col).font = styles.Font(bold=True, color="FF0000")
    wb.save(r'C:\Users\Julio\Desktop\arquivo_editado.xlsx')
